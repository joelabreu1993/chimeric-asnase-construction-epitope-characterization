import os
import glob
import pandas as pd
from openpyxl.styles import Alignment, Font # Para formatar as células no Excel

def parse_rnk_file_to_dataframe(filepath):
    """
    Lê um arquivo .rnk, remove moléculas duplicadas, e o transforma em
    um DataFrame com a coluna 'Mol No' como índice para buscas rápidas.
    """
    column_names = [
        'Mol No', 'Score', 'S(PLP)', 'S(hbond)', 'S(cho)', 'S(metal)',
        'DE(clash)', 'DE(tors)', 'intcor', 'time'
    ]
    try:
        df = pd.read_csv(filepath, skiprows=4, sep='\s+', names=column_names)
        df['Mol No'] = pd.to_numeric(df['Mol No'], errors='coerce')
        df.dropna(subset=['Mol No'], inplace=True)
        df['Mol No'] = df['Mol No'].astype(int)
        df.drop_duplicates(subset='Mol No', keep='first', inplace=True)
        df.set_index('Mol No', inplace=True)
        return df
    except FileNotFoundError:
        print(f"AVISO: O arquivo {filepath} não foi encontrado e será ignorado.")
        return None
    except Exception as e:
        print(f"Ocorreu um erro ao ler o arquivo {filepath}: {e}")
        return None

def main():
    """
    Função principal que encontra e processa múltiplas pastas 'Variante',
    agregando os resultados em um único arquivo Excel formatado.
    """
    # --- PONTO DE EDIÇÃO MANUAL ---
    caminho_monomeros = "../monomeros" # <--- EDITE ESTA LINHA, SE NECESSÁRIO
    # -----------------------------

    if not os.path.isdir(caminho_monomeros):
        print(f"ERRO: A pasta de monômeros não foi encontrada em: '{caminho_monomeros}'")
        return

    padrao_busca = os.path.join(caminho_monomeros, "Variante*_*_monomer")
    lista_pastas_variante = sorted(glob.glob(padrao_busca))

    if not lista_pastas_variante:
        print(f"ERRO: Nenhuma pasta 'Variante...' foi encontrada dentro de '{caminho_monomeros}'")
        return

    print(f"Encontradas {len(lista_pastas_variante)} pastas de variantes para processar...")

    df_agregado = pd.DataFrame({'Pose': list(range(1, 201))})
    df_agregado.set_index('Pose', inplace=True)

    for caminho_variante in lista_pastas_variante:
        nome_da_pasta = os.path.basename(caminho_variante)
        print(f"Processando: {nome_da_pasta}")

        path_asn = os.path.join(caminho_variante, 'L-Asn_m1', 'L-Asn_m1.rnk')
        path_gln = os.path.join(caminho_variante, 'L-Gln_m1', 'L-Gln_m1.rnk')

        df_asparagina = parse_rnk_file_to_dataframe(path_asn)
        df_glutamina = parse_rnk_file_to_dataframe(path_gln)

        poses = list(range(1, 201))
        scores_asn = df_asparagina.reindex(poses)['Score'] if df_asparagina is not None else pd.Series(index=poses, dtype=float)
        scores_gln = df_glutamina.reindex(poses)['Score'] if df_glutamina is not None else pd.Series(index=poses, dtype=float)

        df_agregado[f'{nome_da_pasta}_Asn'] = scores_asn.values
        df_agregado[f'{nome_da_pasta}_Gln'] = scores_gln.values
        df_agregado[f'spacer_{nome_da_pasta}'] = ''

    output_filename_excel = 'resultados_agregados.xlsx'
    print(f"\nCriando o arquivo Excel: {output_filename_excel}")

    with pd.ExcelWriter(output_filename_excel, engine='openpyxl') as writer:
        # --- CORREÇÃO DO ERRO ---
        # Alterado 'startrow=1' para 'startrow=2' para que os dados comecem na linha 3,
        # deixando as linhas 1 e 2 livres para o cabeçalho e evitando a sobrescrita.
        df_agregado.reset_index().to_excel(writer, sheet_name='Resultados', startrow=2, header=False, index=False)

        workbook = writer.book
        worksheet = writer.sheets['Resultados']

        header_font = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center')

        cell_pose = worksheet['A1']
        cell_pose.value = "Pose"
        cell_pose.font = header_font
        cell_pose.alignment = center_align
        worksheet.merge_cells('A1:A2')
        
        for i, caminho_variante in enumerate(lista_pastas_variante):
            nome_da_pasta = os.path.basename(caminho_variante)
            start_col = 2 + (i * 3)
            end_col = start_col + 1
            
            cell_header = worksheet.cell(row=1, column=start_col)
            cell_header.value = nome_da_pasta
            cell_header.font = header_font
            cell_header.alignment = center_align
            worksheet.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

            worksheet.cell(row=2, column=start_col).value = "Asparagina"
            worksheet.cell(row=2, column=end_col).value = "Glutamina"

    print("\nProcesso concluído com sucesso!")
    print(f"O arquivo '{output_filename_excel}' foi criado em: {os.path.abspath(output_filename_excel)}")

if __name__ == "__main__":
    main()